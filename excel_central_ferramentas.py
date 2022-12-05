import openpyxl
from classes_principais import *


### IMPLEMENTAÇÃO DE EXCEL ###


class excel:
# Carrega a planilha se ela já existir, senão cria um arquivo novo #
    def __init__(self, aba):
        self.arquivo_xls = "arquivo_central_ferramentas.xlsx"
        try:
            self.wb = openpyxl.load_workbook(self.arquivo_xls)
            self.sheet = self.wb[aba]
        except:
            self.wb = openpyxl.Workbook()
            self.criar_excel_ferramentas ()
            self.criar_excel_tecnicos ()
            self.wb.save (self.arquivo_xls)
            self.sheet = self.wb[aba]

# Exclui a linha da planilha, da aba, que o usuario escolher #
    def excluir_dados(self, id_item):
        for linha in self.sheet.iter_rows():
            for coluna in linha:
                valor_id = coluna.value
                if valor_id == id_item:
                    self.sheet.delete_rows(linha[0].row, 1)
                else:
                    break
        self.wb.save(self.arquivo_xls)

# criar a linhas e colunas no excel - FERRAMENTAS #        
    def criar_excel_ferramentas(self):
        self.wb.create_sheet('Aba Ferramentas')
        self.sheet = self.wb['Aba Ferramentas']
        self.sheet.column_dimensions['A'].width = 15
        self.sheet.column_dimensions['B'].width = 30
        self.sheet.column_dimensions['C'].width = 30
        self.sheet.column_dimensions['D'].width = 15
        self.sheet.column_dimensions['E'].width = 15
        self.sheet.column_dimensions['F'].width = 20
        self.sheet.column_dimensions['G'].width = 20
        self.sheet.column_dimensions['H'].width = 20
        self.sheet.column_dimensions['I'].width = 20
        
        self.sheet.cell(row=1, column=1).value = "ID da ferramenta"
        self.sheet.cell(row=1, column=2).value = "Descrição da ferramenta"
        self.sheet.cell(row=1, column=3).value = "Fabricante"
        self.sheet.cell(row=1, column=4).value = "Voltagem "
        self.sheet.cell(row=1, column=5).value = "Part Number"
        self.sheet.cell(row=1, column=6).value = "Tamanho"
        self.sheet.cell(row=1, column=7).value = "Tipo de Ferramenta"
        self.sheet.cell(row=1, column=8).value = "Material da Ferramenta"
        self.sheet.cell(row=1, column=9).value = "Máximo de Reserva (hrs)"

 # salvar informações no excel #
    def salvar_ferramentas (self, nova_ferramenta):
        linha_atual = self.sheet.max_row 
        id = self.sheet.cell(row=linha_atual, column=1).value 
        num_id = 0
        if linha_atual > 1:
            num_id = int(id)
        self.sheet.cell(row=linha_atual + 1, column=1).value = num_id + 1
        self.sheet.cell(row=linha_atual + 1, column=2).value = nova_ferramenta.getDescricao_ferramenta()
        self.sheet.cell(row=linha_atual + 1, column=3).value = nova_ferramenta.getFabricante()
        self.sheet.cell(row=linha_atual + 1, column=4).value = nova_ferramenta.getVoltagem()
        self.sheet.cell(row=linha_atual + 1, column=5).value = nova_ferramenta.getPart_number()
        self.sheet.cell(row=linha_atual + 1, column=6).value = nova_ferramenta.getTamanho()
        self.sheet.cell(row=linha_atual + 1, column=7).value = nova_ferramenta.getTipo()
        self.sheet.cell(row=linha_atual + 1, column=8).value = nova_ferramenta.getMaterial()
        self.sheet.cell(row=linha_atual + 1, column=9).value = nova_ferramenta.getMaximo_reserva()
        self.wb.save(self.arquivo_xls)

       
# criar a linhas e colunas no excel - TECNICO#
    def criar_excel_tecnicos(self):
        self.wb.create_sheet('Aba Técnicos')
        self.sheet = self.wb['Aba Técnicos']
        self.sheet.column_dimensions['A'].width = 15
        self.sheet.column_dimensions['B'].width = 20
        self.sheet.column_dimensions['C'].width = 30
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 20
        self.sheet.column_dimensions['G'].width = 40


        self.sheet.cell(row=1, column=1).value = "ID Técnico"
        self.sheet.cell(row=1, column=2).value = "CPF"
        self.sheet.cell(row=1, column=3).value = "Nome Completo"
        self.sheet.cell(row=1, column=4).value = "Telefone"
        self.sheet.cell(row=1, column=5).value = "Turno"
        self.sheet.cell(row=1, column=6).value = "Equipe"
        self.sheet.cell(row=1, column=7).value = "Email"

# salvar informações no excel #
  
    def salvar_tecnicos (self, novo_tecnico):
        linha_atual = self.sheet.max_row 
        id = self.sheet.cell(row=linha_atual, column=1).value 
        num_id = 0
        if linha_atual > 1:
            num_id = int(id)
        self.sheet.cell(row=linha_atual + 1, column=1).value = num_id + 1
        self.sheet.cell(row=linha_atual + 1, column=2).value = novo_tecnico.getCpf()
        self.sheet.cell(row=linha_atual + 1, column=3).value = novo_tecnico.getNome()
        self.sheet.cell(row=linha_atual + 1, column=4).value = novo_tecnico.getTelefone()
        self.sheet.cell(row=linha_atual + 1, column=5).value = novo_tecnico.getTurno()
        self.sheet.cell(row=linha_atual + 1, column=6).value = novo_tecnico.getEquipe()
        self.sheet.cell(row=linha_atual + 1, column=7).value = novo_tecnico.getEmail()
        self.wb.save(self.arquivo_xls)

    

    